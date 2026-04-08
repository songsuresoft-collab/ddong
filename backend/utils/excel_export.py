import io
from openpyxl import Workbook
from fastapi.responses import StreamingResponse

def export_dict_to_excel(data: dict, sheet_title: str = "Export") -> StreamingResponse:
    """
    Converts a dictionary (e.g., meeting minutes or weekly report) into an Excel file stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    # Simple key-value mapping for the dictionary layout
    row = 1
    for key, value in data.items():
        ws.cell(row=row, column=1, value=str(key).replace("_", " ").title())
        
        if isinstance(value, list):
            ws.cell(row=row, column=2, value="\n".join(str(v) for v in value))
        elif isinstance(value, dict):
            ws.cell(row=row, column=2, value=str(value))
        else:
            ws.cell(row=row, column=2, value=str(value))
            
        row += 1
        
    # Basic Formatting
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 100
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{sheet_title.lower().replace(" ", "_")}.xlsx"'}
    )
